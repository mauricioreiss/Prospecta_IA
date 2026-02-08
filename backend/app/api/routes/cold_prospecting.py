"""Cold Prospecting API Routes - SPIN Selling via WhatsApp (isolated from reactivation)"""
import csv
import io
import json
import uuid
import asyncio
from datetime import datetime, timedelta
from typing import Optional, List, Set
from fastapi import APIRouter, File, UploadFile, HTTPException, Form, BackgroundTasks

from backend.app.integrations.n8n import send_whatsapp_duas_mensagens
from backend.app.integrations.supabase import get_supabase_client
from backend.app.config import settings

# Lazy import for AI Responder context
_save_lead_context = None

async def _save_context(phone: str, name: str, notes: str, company: str, campaign_id: str = None):
    global _save_lead_context
    if _save_lead_context is None:
        from backend.app.api.routes.ai_responder import save_lead_context
        _save_lead_context = save_lead_context
    return await _save_lead_context(phone, name, notes, company, campaign_id)

# Excel support
try:
    from openpyxl import load_workbook
    XLSX_SUPPORTED = True
except ImportError:
    XLSX_SUPPORTED = False

router = APIRouter(prefix="/cold-prospecting", tags=["cold-prospecting"])

# =====================================================
# SPIN Selling Message Templates
# =====================================================
SPIN_MSG_1 = "Oi {nome}, tudo bem? João aqui."
SPIN_MSG_2 = "Vi a {empresa} no Google e fiquei com uma dúvida sobre a frota de vocês."

# =====================================================
# Helpers
# =====================================================

def format_spin_messages(name: str, company: str) -> tuple:
    """Format SPIN messages with lead data"""
    nome = name if name and name != "Amigo" else "tudo bem"
    empresa = company if company else "sua empresa"

    msg1 = SPIN_MSG_1.replace("{nome}", nome)
    msg2 = SPIN_MSG_2.replace("{empresa}", empresa)
    return msg1, msg2


def clean_phone(phone: str) -> Optional[str]:
    """Clean phone to 11-digit Brazilian format. Returns None if invalid."""
    phone_clean = ''.join(c for c in phone if c.isdigit())
    if len(phone_clean) < 10:
        return None
    if phone_clean.startswith('55'):
        phone_clean = phone_clean[2:]
    if len(phone_clean) > 11:
        phone_clean = phone_clean[:11]
    return phone_clean


def parse_cold_csv(content: str) -> dict:
    """Parse CSV for cold prospecting. Columns: Nome, Telefone, Empresa."""
    leads = []
    skipped_no_phone = 0
    seen_phones: Set[str] = set()
    duplicates = []

    first_line = content.split('\n')[0]
    delimiter = ',' if ',' in first_line else ';' if ';' in first_line else '\t'
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)

    name_cols = [
        'name', 'nome', 'NAME', 'Nome', 'NOME',
        'contato', 'Contato', 'CONTATO',
        'Dono(s)', 'dono(s)', 'Dono', 'dono', 'DONO',
        'responsavel', 'Responsavel'
    ]
    phone_cols = [
        'phone', 'telefone', 'PHONE', 'Telefone', 'TELEFONE',
        'tel', 'Tel', 'TEL',
        'celular', 'Celular', 'CELULAR',
        'whatsapp', 'WhatsApp', 'WHATSAPP',
        'fone', 'Fone'
    ]
    company_cols = [
        'company', 'empresa', 'COMPANY', 'Empresa', 'EMPRESA',
        'razao_social', 'Razao_Social', 'RAZAO_SOCIAL'
    ]

    def find_column(row: dict, possible_names: list) -> str:
        for col in possible_names:
            if col in row and row[col]:
                return str(row[col]).strip()
        return ""

    for row in reader:
        name = find_column(row, name_cols)
        phone = find_column(row, phone_cols)
        company = find_column(row, company_cols)

        if not phone:
            skipped_no_phone += 1
            continue

        phone_clean = clean_phone(phone)
        if not phone_clean:
            skipped_no_phone += 1
            continue

        if phone_clean in seen_phones:
            duplicates.append({'name': name, 'phone': phone_clean})
            continue

        seen_phones.add(phone_clean)
        leads.append({
            'name': name or 'Amigo',
            'phone': phone_clean,
            'company': company or ''
        })

    return {
        'leads': leads,
        'skipped_no_phone': skipped_no_phone,
        'duplicates': duplicates,
        'sheets_processed': [],
        'skipped_sheets': []
    }


def parse_cold_xlsx(file_bytes: bytes) -> dict:
    """Parse XLSX for cold prospecting. Columns: Nome, Telefone, Empresa."""
    if not XLSX_SUPPORTED:
        raise ValueError("openpyxl nao instalado - suporte a xlsx indisponivel")

    leads = []
    skipped_no_phone = 0
    seen_phones: Set[str] = set()
    duplicates = []
    sheets_processed = []

    name_cols = [
        'name', 'nome', 'NAME', 'Nome', 'NOME',
        'contato', 'Contato', 'CONTATO',
        'Dono(s)', 'dono(s)', 'Dono', 'dono', 'DONO',
        'responsavel', 'Responsavel'
    ]
    phone_cols = [
        'phone', 'telefone', 'PHONE', 'Telefone', 'TELEFONE',
        'tel', 'Tel', 'TEL',
        'celular', 'Celular', 'CELULAR',
        'whatsapp', 'WhatsApp', 'WHATSAPP',
        'fone', 'Fone'
    ]
    company_cols = [
        'company', 'empresa', 'COMPANY', 'Empresa', 'EMPRESA',
        'razao_social', 'Razao_Social', 'RAZAO_SOCIAL'
    ]

    def find_col_index(headers: list, possible_names: list) -> int:
        for i, h in enumerate(headers):
            if h in possible_names:
                return i
        return -1

    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    header_keywords = ['telefone', 'phone', 'dono', 'empresa', 'nome', 'contato']
    summary_keywords = ['total', 'resumo', 'dashboard', 'closer', 'kpi', 'meta', 'consolidado']
    skipped_sheets = []

    for sheet_name in wb.sheetnames:
        sheet_name_lower = sheet_name.lower()
        if any(kw in sheet_name_lower for kw in summary_keywords):
            skipped_sheets.append({'name': sheet_name, 'reason': 'Aba de resumo/dashboard'})
            continue

        ws = wb[sheet_name]
        sheet_leads_count = 0
        headers = []

        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
            row_values = [str(cell.value or '').strip().lower() for cell in row]
            if any(keyword in val for val in row_values for keyword in header_keywords):
                headers = [str(cell.value or '').strip() for cell in row]
                header_row_num = row_num
                break

        if not headers:
            continue

        name_idx = find_col_index(headers, name_cols)
        phone_idx = find_col_index(headers, phone_cols)
        company_idx = find_col_index(headers, company_cols)

        if phone_idx < 0:
            continue

        for row in ws.iter_rows(min_row=header_row_num + 1):
            values = [str(cell.value or '').strip() for cell in row]
            if not any(values):
                continue

            name = values[name_idx] if name_idx >= 0 and name_idx < len(values) else ''
            phone = values[phone_idx] if phone_idx >= 0 and phone_idx < len(values) else ''
            company = values[company_idx] if company_idx >= 0 and company_idx < len(values) else ''

            if not phone:
                skipped_no_phone += 1
                continue

            phone_clean = clean_phone(phone)
            if not phone_clean:
                skipped_no_phone += 1
                continue

            if phone_clean in seen_phones:
                duplicates.append({'name': name, 'phone': phone_clean, 'sheet': sheet_name})
                continue

            seen_phones.add(phone_clean)
            sheet_leads_count += 1
            leads.append({
                'name': name or 'Amigo',
                'phone': phone_clean,
                'company': company or '',
                'source_sheet': sheet_name
            })

        if sheet_leads_count > 0:
            sheets_processed.append({'name': sheet_name, 'leads': sheet_leads_count})

    wb.close()

    return {
        'leads': leads,
        'skipped_no_phone': skipped_no_phone,
        'duplicates': duplicates,
        'sheets_processed': sheets_processed,
        'skipped_sheets': skipped_sheets
    }


# =====================================================
# Core: Send function (isolated from reactivation)
# =====================================================

async def enviar_prospeccao_fria(lead_data: dict) -> dict:
    """
    Send SPIN Selling cold prospecting messages for a single lead.
    Uses the same n8n webhook but with SPIN templates and tag 'prospeccao_fria_spin'.
    """
    name = lead_data.get('name', 'Amigo')
    phone = lead_data.get('phone', '')
    company = lead_data.get('company', '')

    msg1, msg2 = format_spin_messages(name, company)

    result = await send_whatsapp_duas_mensagens(
        phone=phone,
        msg1=msg1,
        msg2=msg2,
        tag_campanha="prospeccao_fria_spin",
        inbox_id=366
    )

    return result


async def log_cold_send(phone: str, name: str, company: str, campaign_id: str, status: str = 'sent', error: str = None):
    """Log a cold prospecting send to the database"""
    try:
        client = get_supabase_client()
        client.table('reactivation_log').insert({
            'phone': phone,
            'name': name,
            'company': company,
            'campaign_id': campaign_id,
            'status': status,
            'error': error,
            'sent_at': datetime.now().isoformat()
        }).execute()
    except Exception as e:
        print(f"Warning: Could not log cold send for {phone}: {e}")


async def check_already_contacted_cold(phones: List[str], days: int = 30) -> Set[str]:
    """Check which phones were already contacted recently"""
    already_contacted = set()
    try:
        client = get_supabase_client()
        threshold_date = (datetime.now() - timedelta(days=days)).isoformat()
        try:
            result = client.table('reactivation_log').select(
                "phone"
            ).in_(
                "phone", phones
            ).gte(
                "sent_at", threshold_date
            ).execute()
            if result.data:
                for row in result.data:
                    if row.get('phone'):
                        already_contacted.add(row['phone'])
        except Exception:
            pass
    except Exception as e:
        print(f"Warning: Could not check contacted phones: {e}")
    return already_contacted


async def send_cold_messages_sequentially(
    leads: List[dict],
    campaign_id: str,
    delay_seconds: int
):
    """
    Send SPIN cold prospecting messages sequentially.
    Delay only between different leads.
    """
    sent = 0
    failed = 0

    for i, lead in enumerate(leads):
        phone = lead['phone']
        name = lead.get('name', '')
        company = lead.get('company', '')

        try:
            result = await enviar_prospeccao_fria(lead)

            if result['success']:
                print(f"[{campaign_id}] {i+1}/{len(leads)} - OK {phone} (SPIN 2 msgs)")
                await _save_context(
                    phone=phone,
                    name=name,
                    notes=f"Prospeccao fria SPIN - empresa: {company}",
                    company=company,
                    campaign_id=campaign_id
                )
                await log_cold_send(phone, name, company, campaign_id, status='sent')
                sent += 1
            else:
                await log_cold_send(phone, name, company, campaign_id, status='failed', error=result.get('error'))
                failed += 1
                print(f"[{campaign_id}] {i+1}/{len(leads)} - FALHOU {phone}: {result.get('error')}")

        except Exception as e:
            await log_cold_send(phone, name, company, campaign_id, status='failed', error=str(e))
            failed += 1
            print(f"[{campaign_id}] {i+1}/{len(leads)} - ERRO {phone}: {e}")

        if i < len(leads) - 1:
            await asyncio.sleep(delay_seconds)

    print(f"[{campaign_id}] COMPLETO - Enviados: {sent}, Falhos: {failed}")


# =====================================================
# API Endpoints
# =====================================================

@router.post("/preview")
async def preview_cold_prospecting(
    file: UploadFile = File(...),
    check_days: int = Form(30)
):
    """
    Upload CSV/XLSX and preview cold leads WITHOUT sending.
    Returns parsed leads with SPIN message preview and safety report.
    """
    filename = file.filename.lower() if file.filename else ''

    if not filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Formato invalido. Use CSV ou XLSX")

    try:
        content = await file.read()

        if filename.endswith('.csv'):
            text = None
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                raise HTTPException(status_code=400, detail="Erro ao ler CSV - encoding invalido")
            parsed = parse_cold_csv(text)
        else:
            if not XLSX_SUPPORTED:
                raise HTTPException(status_code=400, detail="Suporte a Excel nao disponivel. Instale openpyxl")
            parsed = parse_cold_xlsx(content)

        leads = parsed['leads']
        if not leads:
            raise HTTPException(status_code=400, detail="Nenhum lead encontrado. Verifique colunas: Nome, Telefone, Empresa")

        # Check already contacted
        all_phones = [l['phone'] for l in leads]
        already_contacted = await check_already_contacted_cold(all_phones, days=check_days)

        sendable = [l for l in leads if l['phone'] not in already_contacted]
        skipped_contacted = [l for l in leads if l['phone'] in already_contacted]

        if not sendable:
            raise HTTPException(status_code=400, detail="Todos os leads ja foram contatados recentemente")

        # Preview first 5 with SPIN messages
        preview = []
        for lead in sendable[:5]:
            msg1, msg2 = format_spin_messages(lead['name'], lead['company'])
            preview.append({
                **lead,
                'preview_msg1': msg1,
                'preview_msg2': msg2
            })

        return {
            'total_leads': len(sendable),
            'preview': preview,
            'all_leads': sendable,
            'safety_report': {
                'total_in_file': len(leads) + parsed['skipped_no_phone'] + len(parsed['duplicates']),
                'sendable': len(sendable),
                'already_contacted': len(skipped_contacted),
                'already_contacted_list': skipped_contacted[:10],
                'duplicates_removed': len(parsed['duplicates']),
                'duplicates_list': parsed['duplicates'][:10],
                'no_phone': parsed['skipped_no_phone'],
                'sheets_processed': parsed.get('sheets_processed', []),
                'sheets_skipped': parsed.get('skipped_sheets', [])
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")


@router.post("/send")
async def send_cold_prospecting(
    background_tasks: BackgroundTasks,
    leads: str = Form(...),
    delay_seconds: int = Form(45)
):
    """
    Send SPIN Selling messages to previewed leads.
    Accepts JSON string of leads (from preview response).
    """
    try:
        leads_list = json.loads(leads)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Formato de leads invalido")

    if not leads_list:
        raise HTTPException(status_code=400, detail="Lista de leads vazia")

    campaign_id = f"cold_spin_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"

    background_tasks.add_task(
        send_cold_messages_sequentially,
        leads=leads_list,
        campaign_id=campaign_id,
        delay_seconds=delay_seconds
    )

    return {
        'status': 'sending',
        'campaign_id': campaign_id,
        'total': len(leads_list),
        'messages_per_lead': 2,
        'delay_seconds': delay_seconds,
        'estimated_time_minutes': (len(leads_list) * delay_seconds) // 60,
        'message': f'Prospeccao fria SPIN iniciada: {len(leads_list)} leads (2 msgs cada)'
    }


@router.get("/campaign/{campaign_id}")
async def get_cold_campaign_progress(campaign_id: str):
    """Get progress of a cold prospecting campaign."""
    try:
        client = get_supabase_client()
        result = client.table('reactivation_log').select(
            "status"
        ).eq(
            "campaign_id", campaign_id
        ).execute()

        if not result.data:
            return {
                'campaign_id': campaign_id,
                'found': False,
                'message': 'Campanha nao encontrada ou ainda nao iniciou'
            }

        sent = sum(1 for r in result.data if r['status'] == 'sent')
        failed = sum(1 for r in result.data if r['status'] == 'failed')
        total = len(result.data)

        return {
            'campaign_id': campaign_id,
            'found': True,
            'total_processed': total,
            'sent': sent,
            'failed': failed,
            'progress_percent': round((total / max(total, 1)) * 100, 1)
        }
    except Exception as e:
        return {
            'campaign_id': campaign_id,
            'found': False,
            'error': str(e)
        }


@router.get("/template")
def get_spin_template():
    """Get the SPIN Selling message templates"""
    example_msg1, example_msg2 = format_spin_messages("Carlos", "Auto Pecas Silva")
    return {
        'spin_msg1': SPIN_MSG_1,
        'spin_msg2': SPIN_MSG_2,
        'example': {
            'msg1': example_msg1,
            'msg2': example_msg2
        },
        'info': 'SPIN Selling: 2 mensagens de isca (abertura + curiosidade sobre frota). Inbox 366.'
    }
