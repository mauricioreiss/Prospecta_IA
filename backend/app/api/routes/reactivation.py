"""Reactivation API Routes - CSV/XLSX upload and bulk WhatsApp for old leads"""
import csv
import io
from datetime import datetime, timedelta
from typing import Optional, List, Set
from fastapi import APIRouter, File, UploadFile, HTTPException, Form, BackgroundTasks
from pydantic import BaseModel

import asyncio
from backend.app.integrations.n8n import send_whatsapp_single, send_whatsapp_duas_mensagens
from backend.app.integrations.supabase import get_supabase_client
from backend.app.config import settings

# Import for AI Responder context storage
# Lazy import to avoid circular dependency
_save_lead_context = None

async def save_lead_context_wrapper(phone: str, name: str, notes: str, company: str, campaign_id: str = None):
    """Wrapper to lazily import and call save_lead_context"""
    global _save_lead_context
    if _save_lead_context is None:
        from backend.app.api.routes.ai_responder import save_lead_context
        _save_lead_context = save_lead_context
    return await _save_lead_context(phone, name, notes, company, campaign_id)

# Excel support
try:
    from openpyxl import load_workbook
    XLSX_SUPPORTED = True
except ImportError:
    XLSX_SUPPORTED = False

router = APIRouter(prefix="/reactivation", tags=["reactivation"])

# Status que NUNCA devem receber mensagem (cliente já fechou)
BLOCKED_STATUS = [
    'fechado', 'fechou', 'closed', 'won', 'ganho', 'vendido',
    'cliente', 'ativo', 'contrato', 'assinado'
]

# Message 1 - Opening + Question (generic pain point)
DEFAULT_MSG1_TEMPLATE = """Fala, [NAME]! Tudo bem por aí? João aqui.

Estava revisando aqui as anotações que meu time fez sobre sua empresa. Vi que você estava dependendo de indicações para fazer novas locações. Conseguiu resolver esse problema?"""

# Message 2 - Value proposition + CTA + Link (all together)
DEFAULT_MSG2_TEMPLATE = """Pergunto porque amanhã teremos uma aula gratuita online com um cliente nosso que saiu de R$ 65.000 para R$ 150.000 sem depender de indicação! Quero te convidar para esse evento. Faz sentido pra você?

Se sim, entre no link abaixo:

https://webnario2.oduoassessoria.com.br/?utm_source=kpis_joao&utm_medium=wpp&utm_campaign=1x1&utm_content=msg1&utm_term=lead_joao"""

# Keep backward compatibility
DEFAULT_REACTIVATION_TEMPLATE = DEFAULT_MSG1_TEMPLATE
DEFAULT_LINK_TEMPLATE = DEFAULT_MSG2_TEMPLATE


class ReactivationLead(BaseModel):
    """Single lead from CSV for reactivation"""
    name: str
    phone: str
    notes: Optional[str] = ""
    company: Optional[str] = ""


class ReactivationRequest(BaseModel):
    """Bulk reactivation request"""
    leads: List[ReactivationLead]
    message_template: str = DEFAULT_REACTIVATION_TEMPLATE
    delay_seconds: int = 45  # Delay between messages to avoid spam


class ReactivationResult(BaseModel):
    """Result of reactivation campaign"""
    total: int
    queued: int
    failed: int
    errors: List[str]


def parse_csv_content(content: str) -> dict:
    """
    Parse CSV content and extract leads.

    Columns from user's KPI spreadsheet:
    - Dono(s) = NAME
    - Empresa = COMPANY
    - Telefone = PHONE
    - Resultado = STATUS (skip if FECHADO)
    - Resumo = NOTES

    Returns dict with:
    - leads: valid leads to contact
    - skipped_fechado: leads skipped because status is FECHADO
    - skipped_no_phone: leads without phone
    - duplicates: duplicate phone numbers removed
    """
    leads = []
    skipped_fechado = []
    skipped_no_phone = 0
    seen_phones: Set[str] = set()
    duplicates = []

    # Try to detect delimiter
    first_line = content.split('\n')[0]
    delimiter = ',' if ',' in first_line else ';' if ';' in first_line else '\t'

    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)

    # Normalize column names - including user's KPI spreadsheet format
    name_cols = [
        'name', 'nome', 'NAME', 'Nome', 'NOME',
        'contato', 'Contato', 'CONTATO',
        'Dono(s)', 'dono(s)', 'Dono', 'dono', 'DONO',  # User's spreadsheet
        'responsavel', 'Responsavel'
    ]
    phone_cols = [
        'phone', 'telefone', 'PHONE', 'Telefone', 'TELEFONE',
        'tel', 'Tel', 'TEL',
        'celular', 'Celular', 'CELULAR',
        'whatsapp', 'WhatsApp', 'WHATSAPP',
        'fone', 'Fone'
    ]
    notes_cols = [
        'notes', 'notas', 'NOTES', 'Notas', 'NOTAS',
        'observacao', 'Observacao', 'OBSERVACAO',
        'obs', 'Obs', 'OBS',
        'dificuldade', 'Dificuldade', 'problema', 'Problema',
        'resumo', 'Resumo', 'RESUMO'  # User's spreadsheet
    ]
    company_cols = [
        'company', 'empresa', 'COMPANY', 'Empresa', 'EMPRESA',
        'razao_social', 'Razao_Social', 'RAZAO_SOCIAL'
    ]
    status_cols = [
        'resultado', 'Resultado', 'RESULTADO',  # User's spreadsheet
        'status', 'Status', 'STATUS',
        'fase', 'Fase', 'FASE',
        'situacao', 'Situacao', 'SITUACAO'
    ]

    def find_column(row: dict, possible_names: list) -> str:
        for col in possible_names:
            if col in row and row[col]:
                return str(row[col]).strip()
        return ""

    for row in reader:
        name = find_column(row, name_cols)
        phone = find_column(row, phone_cols)
        notes = find_column(row, notes_cols)
        company = find_column(row, company_cols)
        status = find_column(row, status_cols)

        if not phone:
            skipped_no_phone += 1
            continue

        # Clean phone number - keep only digits
        phone_clean = ''.join(c for c in phone if c.isdigit())
        if len(phone_clean) < 10:
            skipped_no_phone += 1
            continue

        # FORCE 11 digits (DDD + 9 digit phone)
        # Remove 55 from start if present
        if phone_clean.startswith('55'):
            phone_clean = phone_clean[2:]
        # If still more than 11 digits, take only first 11
        if len(phone_clean) > 11:
            phone_clean = phone_clean[:11]

        # CHECK 1: Skip if status is FECHADO (client already bought)
        status_lower = status.lower() if status else ''
        if any(blocked in status_lower for blocked in BLOCKED_STATUS):
            skipped_fechado.append({
                'name': name or 'Sem nome',
                'company': company,
                'phone': phone_clean,
                'status': status
            })
            continue

        # CHECK 2: Skip duplicates (same phone number)
        if phone_clean in seen_phones:
            duplicates.append({
                'name': name,
                'phone': phone_clean
            })
            continue

        seen_phones.add(phone_clean)

        leads.append({
            'name': name or 'Amigo',
            'phone': phone_clean,
            'notes': notes or 'queria crescer o negócio',
            'company': company or '',
            'original_status': status
        })

    return {
        'leads': leads,
        'skipped_fechado': skipped_fechado,
        'skipped_no_phone': skipped_no_phone,
        'duplicates': duplicates,
        'sheets_processed': [],  # CSV doesn't have sheets
        'skipped_sheets': []  # CSV doesn't have sheets
    }


def parse_xlsx_content(file_bytes: bytes) -> dict:
    """
    Parse XLSX content and extract leads from ALL sheets/pages.
    Same logic as CSV parser but for Excel files with multiple sheets.

    Processes ALL sheets in the workbook and deduplicates across all of them.
    """
    if not XLSX_SUPPORTED:
        raise ValueError("openpyxl nao instalado - suporte a xlsx indisponivel")

    leads = []
    skipped_fechado = []
    skipped_no_phone = 0
    seen_phones: Set[str] = set()  # Deduplicate across ALL sheets
    duplicates = []
    sheets_processed = []

    # Column mapping - same as CSV
    name_cols = [
        'name', 'nome', 'NAME', 'Nome', 'NOME',
        'contato', 'Contato', 'CONTATO',
        'Dono(s)', 'dono(s)', 'Dono', 'dono', 'DONO',
        'responsavel', 'Responsavel'
    ]
    phone_cols = [
        'phone', 'telefone', 'PHONE', 'Telefone', 'TELEFONE',
        'tel', 'Tel', 'TEL',
        'celular', 'Celular', 'CELULAR',
        'whatsapp', 'WhatsApp', 'WHATSAPP',
        'fone', 'Fone'
    ]
    notes_cols = [
        'notes', 'notas', 'NOTES', 'Notas', 'NOTAS',
        'observacao', 'Observacao', 'OBSERVACAO',
        'obs', 'Obs', 'OBS',
        'dificuldade', 'Dificuldade', 'problema', 'Problema',
        'resumo', 'Resumo', 'RESUMO'
    ]
    company_cols = [
        'company', 'empresa', 'COMPANY', 'Empresa', 'EMPRESA',
        'razao_social', 'Razao_Social', 'RAZAO_SOCIAL'
    ]
    status_cols = [
        'resultado', 'Resultado', 'RESULTADO',
        'status', 'Status', 'STATUS',
        'fase', 'Fase', 'FASE',
        'situacao', 'Situacao', 'SITUACAO'
    ]

    def find_col_index(headers: list, possible_names: list) -> int:
        for i, h in enumerate(headers):
            if h in possible_names:
                return i
        return -1

    # Load workbook from bytes
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)

    # Keywords to identify the header row
    header_keywords = ['telefone', 'phone', 'dono', 'empresa', 'nome', 'contato']

    # Keywords that indicate a summary/dashboard sheet (should be skipped)
    summary_keywords = ['total', 'resumo', 'dashboard', 'closer', 'kpi', 'meta', 'consolidado']
    skipped_sheets = []

    # Process ALL sheets in the workbook
    for sheet_name in wb.sheetnames:
        # Skip summary/dashboard sheets
        sheet_name_lower = sheet_name.lower()
        if any(kw in sheet_name_lower for kw in summary_keywords):
            skipped_sheets.append({'name': sheet_name, 'reason': 'Aba de resumo/dashboard'})
            continue

        ws = wb[sheet_name]
        sheet_leads_count = 0

        # Find the header row (may not be row 1 if there's a title)
        header_row_num = 1
        headers = []

        for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=10), start=1):
            row_values = [str(cell.value or '').strip().lower() for cell in row]
            # Check if this row contains header keywords
            if any(keyword in val for val in row_values for keyword in header_keywords):
                headers = [str(cell.value or '').strip() for cell in row]
                header_row_num = row_num
                break

        if not headers:
            continue

        # Find column indices for this sheet
        name_idx = find_col_index(headers, name_cols)
        phone_idx = find_col_index(headers, phone_cols)
        notes_idx = find_col_index(headers, notes_cols)
        company_idx = find_col_index(headers, company_cols)
        status_idx = find_col_index(headers, status_cols)

        # Skip sheet if no phone column found
        if phone_idx < 0:
            continue

        # Process rows (skip header row)
        for row in ws.iter_rows(min_row=header_row_num + 1):
            values = [str(cell.value or '').strip() for cell in row]

            # Skip empty rows
            if not any(values):
                continue

            name = values[name_idx] if name_idx >= 0 and name_idx < len(values) else ''
            phone = values[phone_idx] if phone_idx >= 0 and phone_idx < len(values) else ''
            notes = values[notes_idx] if notes_idx >= 0 and notes_idx < len(values) else ''
            company = values[company_idx] if company_idx >= 0 and company_idx < len(values) else ''
            status = values[status_idx] if status_idx >= 0 and status_idx < len(values) else ''

            if not phone:
                skipped_no_phone += 1
                continue

            # Clean phone number - keep only digits
            phone_clean = ''.join(c for c in phone if c.isdigit())
            if len(phone_clean) < 10:
                skipped_no_phone += 1
                continue

            # FORCE 11 digits (DDD + 9 digit phone)
            # Remove 55 from start if present
            if phone_clean.startswith('55'):
                phone_clean = phone_clean[2:]
            # If still more than 11 digits, take only first 11
            if len(phone_clean) > 11:
                phone_clean = phone_clean[:11]

            # CHECK 1: Skip if status is FECHADO
            status_lower = status.lower() if status else ''
            if any(blocked in status_lower for blocked in BLOCKED_STATUS):
                skipped_fechado.append({
                    'name': name or 'Sem nome',
                    'company': company,
                    'phone': phone_clean,
                    'status': status,
                    'sheet': sheet_name
                })
                continue

            # CHECK 2: Skip duplicates (across ALL sheets)
            if phone_clean in seen_phones:
                duplicates.append({
                    'name': name,
                    'phone': phone_clean,
                    'sheet': sheet_name
                })
                continue

            seen_phones.add(phone_clean)
            sheet_leads_count += 1

            leads.append({
                'name': name or 'Amigo',
                'phone': phone_clean,
                'notes': notes or 'queria crescer o negócio',
                'company': company or '',
                'original_status': status,
                'source_sheet': sheet_name
            })

        if sheet_leads_count > 0:
            sheets_processed.append({'name': sheet_name, 'leads': sheet_leads_count})

    wb.close()

    return {
        'leads': leads,
        'skipped_fechado': skipped_fechado,
        'skipped_no_phone': skipped_no_phone,
        'duplicates': duplicates,
        'sheets_processed': sheets_processed,
        'skipped_sheets': skipped_sheets  # Summary/dashboard sheets that were ignored
    }


def format_message(template: str, lead: dict) -> str:
    """Replace placeholders in template with lead data"""
    message = template
    message = message.replace('[NAME]', lead.get('name', 'Amigo'))
    message = message.replace('[NOTES]', lead.get('notes', 'queria crescer o negócio'))
    message = message.replace('[COMPANY]', lead.get('company', ''))
    message = message.replace('{name}', lead.get('name', 'Amigo'))
    message = message.replace('{notes}', lead.get('notes', 'queria crescer o negócio'))
    message = message.replace('{company}', lead.get('company', ''))
    return message


async def check_already_contacted(phones: List[str], days: int = 30) -> Set[str]:
    """
    Check which phone numbers have already been contacted in the last X days.
    Returns set of phone numbers that should NOT be contacted again.

    Note: This is optional - if database doesn't have the right columns,
    we just return empty set and allow all sends.
    """
    already_contacted = set()

    try:
        client = get_supabase_client()
        threshold_date = (datetime.now() - timedelta(days=days)).isoformat()

        # Try to check reactivation_log table (our own tracking)
        try:
            log_result = client.table('reactivation_log').select(
                "phone"
            ).in_(
                "phone", phones
            ).gte(
                "sent_at", threshold_date
            ).execute()

            if log_result.data:
                for row in log_result.data:
                    if row.get('phone'):
                        already_contacted.add(row['phone'])
        except Exception:
            # Table might not exist yet, that's ok
            pass

    except Exception as e:
        print(f"Warning: Could not check contacted phones (will allow all): {e}")

    return already_contacted


async def log_single_send(phone: str, name: str, company: str, campaign_id: str, status: str = 'sent', error: str = None):
    """Log a single message send to the database"""
    try:
        client = get_supabase_client()
        client.table('reactivation_log').insert({
            'phone': phone,
            'name': name,
            'company': company,
            'campaign_id': campaign_id,
            'status': status,
            'error': error,
            'sent_at': datetime.now().isoformat()
        }).execute()
    except Exception as e:
        print(f"Warning: Could not log send for {phone}: {e}")


async def send_messages_sequentially(
    messages: List[dict],
    campaign_id: str,
    delay_seconds: int,
    tag_campanha: str = "reativacao_oduo"
):
    """
    Send 2 messages per lead (NO delay between them, only between leads):
    - MSG1: Opening + personalized question
    - MSG2: Value proposition + CTA + link

    Delay only happens BETWEEN different leads, not between messages for same lead.
    """
    sent = 0
    failed = 0

    for i, msg in enumerate(messages):
        phone = msg['phone']
        msg1 = msg['msg1']
        msg2 = msg['msg2']
        name = msg.get('name', '')
        company = msg.get('company', '')
        notes = msg.get('notes', '')

        try:
            # Send BOTH messages in ONE call (one conversation, two messages)
            result = await send_whatsapp_duas_mensagens(
                phone=phone,
                msg1=msg1,
                msg2=msg2,
                tag_campanha=tag_campanha
            )

            if result['success']:
                print(f"[{campaign_id}] {i+1}/{len(messages)} - ✓ {phone} (2 msgs)")
                await save_lead_context_wrapper(
                    phone=phone,
                    name=name,
                    notes=notes,
                    company=company,
                    campaign_id=campaign_id
                )
                await log_single_send(phone, name, company, campaign_id, status='sent')
                sent += 1
            else:
                await log_single_send(phone, name, company, campaign_id, status='failed', error=result.get('error'))
                failed += 1
                print(f"[{campaign_id}] {i+1}/{len(messages)} - FALHOU {phone}: {result.get('error')}")

        except Exception as e:
            await log_single_send(phone, name, company, campaign_id, status='failed', error=str(e))
            failed += 1
            print(f"[{campaign_id}] {i+1}/{len(messages)} - ERRO {phone}: {e}")

        # Wait ONLY between different leads (not between messages for same lead)
        if i < len(messages) - 1:
            await asyncio.sleep(delay_seconds)

    print(f"[{campaign_id}] COMPLETO - Enviados: {sent}, Falhos: {failed}")


@router.post("/preview-csv")
@router.post("/preview")
async def preview_file(file: UploadFile = File(...), check_days: int = 30):
    """
    Upload CSV or XLSX and preview the leads that will be contacted.
    Returns parsed leads without sending messages.

    Supported formats: .csv, .xlsx, .xls

    Automatically filters:
    - Leads with FECHADO status (already clients)
    - Duplicate phone numbers
    - Numbers already contacted in the last X days
    """
    filename = file.filename.lower() if file.filename else ''

    # Validate file type
    if not filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Formato invalido. Use CSV ou XLSX (Excel)"
        )

    try:
        content = await file.read()

        # Parse based on file type
        if filename.endswith('.csv'):
            # Try different encodings for CSV
            text = None
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    text = content.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if text is None:
                raise HTTPException(status_code=400, detail="Erro ao ler CSV - encoding invalido")

            parsed = parse_csv_content(text)

        elif filename.endswith(('.xlsx', '.xls')):
            # Parse Excel file
            if not XLSX_SUPPORTED:
                raise HTTPException(
                    status_code=400,
                    detail="Suporte a Excel nao disponivel. Use CSV ou instale openpyxl"
                )
            parsed = parse_xlsx_content(content)
        leads = parsed['leads']
        skipped_fechado = parsed['skipped_fechado']
        skipped_no_phone = parsed['skipped_no_phone']
        duplicates = parsed['duplicates']
        sheets_processed = parsed.get('sheets_processed', [])  # Only for XLSX
        skipped_sheets = parsed.get('skipped_sheets', [])  # Summary sheets ignored

        if not leads and not skipped_fechado:
            raise HTTPException(status_code=400, detail="Nenhum lead encontrado. Verifique se as colunas estao corretas (Dono(s), Empresa, Telefone, Resultado)")

        # Check which phones were already contacted recently
        all_phones = [lead['phone'] for lead in leads]
        already_contacted = await check_already_contacted(all_phones, days=check_days)

        # Separate leads into sendable and already contacted
        sendable_leads = []
        contacted_leads = []

        for lead in leads:
            if lead['phone'] in already_contacted:
                contacted_leads.append({
                    **lead,
                    'reason': f'Ja contatado nos ultimos {check_days} dias'
                })
            else:
                sendable_leads.append(lead)

        # Preview with sample message (first 5 sendable)
        # Using fixed generic pain point, no AI cleaning needed
        preview_leads = []
        for lead in sendable_leads[:5]:
            preview_leads.append({
                **lead,
                'preview_msg1': format_message(DEFAULT_MSG1_TEMPLATE, lead),
                'preview_msg2': format_message(DEFAULT_MSG2_TEMPLATE, lead)
            })

        return {
            # Main data
            'total_leads': len(sendable_leads),
            'preview': preview_leads,
            'all_leads': sendable_leads,
            'default_template': DEFAULT_REACTIVATION_TEMPLATE,

            # Safety report - what was filtered out
            'safety_report': {
                'skipped_fechado': len(skipped_fechado),
                'skipped_fechado_list': skipped_fechado[:10],  # Show first 10
                'already_contacted': len(contacted_leads),
                'already_contacted_list': contacted_leads[:10],
                'duplicates_removed': len(duplicates),
                'duplicates_list': duplicates[:10],  # Show first 10 duplicates
                'no_phone': skipped_no_phone,
                'total_original': len(sendable_leads) + len(skipped_fechado) + len(contacted_leads) + len(duplicates) + skipped_no_phone,
                'sheets_processed': sheets_processed,  # Which Excel sheets were read
                'sheets_skipped': skipped_sheets  # Summary/dashboard sheets ignored
            }
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar CSV: {str(e)}")


@router.post("/send-bulk")
async def send_reactivation_bulk(
    background_tasks: BackgroundTasks,
    leads: str = Form(...),  # JSON string of leads
    msg1_template: str = Form(DEFAULT_MSG1_TEMPLATE),  # Opening + question
    msg2_template: str = Form(DEFAULT_MSG2_TEMPLATE),  # Value prop + CTA + link
    delay_seconds: int = Form(45)
):
    """
    Send reactivation WhatsApp messages to all leads.

    Sends 2 messages per lead (NO delay between them):
    - MSG1: Opening + personalized question
    - MSG2: Value proposition + CTA + link

    Delay only happens between DIFFERENT leads.
    """
    import json
    import uuid

    try:
        leads_list = json.loads(leads)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Formato de leads invalido")

    if not leads_list:
        raise HTTPException(status_code=400, detail="Lista de leads vazia")

    # Generate campaign ID for tracking
    campaign_id = f"reativacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}"

    # Double-check: filter out any FECHADO that might have slipped through
    safe_leads = []
    for lead in leads_list:
        status = lead.get('original_status', '').lower()
        if not any(blocked in status for blocked in BLOCKED_STATUS):
            safe_leads.append(lead)

    if not safe_leads:
        raise HTTPException(status_code=400, detail="Todos os leads foram filtrados (status fechado)")

    # Prepare messages with formatted text (2 messages per lead)
    # Template now uses fixed generic pain point, no need for AI cleaning
    messages_to_send = []
    for lead in safe_leads:
        messages_to_send.append({
            'phone': lead['phone'],
            'msg1': format_message(msg1_template, lead),  # Opening + generic question
            'msg2': format_message(msg2_template, lead),  # Value prop + CTA + link
            'name': lead.get('name', ''),
            'company': lead.get('company', ''),
            'notes': lead.get('notes', '')
        })

    # Start background task to send messages one by one
    # This returns immediately so the frontend doesn't timeout
    background_tasks.add_task(
        send_messages_sequentially,
        messages=messages_to_send,
        campaign_id=campaign_id,
        delay_seconds=delay_seconds,
        tag_campanha="reativacao_oduo"
    )

    return {
        'status': 'sending',
        'campaign_id': campaign_id,
        'total': len(messages_to_send),
        'messages_per_lead': 2,  # 2 messages: opening, then value prop + link
        'delay_seconds': delay_seconds,
        'estimated_time_minutes': (len(messages_to_send) * delay_seconds) // 60,
        'message': f'Iniciando envio de {len(messages_to_send)} leads (2 msgs cada, delay so entre leads)'
    }


@router.get("/campaign/{campaign_id}")
async def get_campaign_progress(campaign_id: str):
    """
    Get the progress of a reactivation campaign.
    Shows how many messages were sent, failed, etc.
    """
    try:
        client = get_supabase_client()

        result = client.table('reactivation_log').select(
            "status"
        ).eq(
            "campaign_id", campaign_id
        ).execute()

        if not result.data:
            return {
                'campaign_id': campaign_id,
                'found': False,
                'message': 'Campanha nao encontrada ou ainda nao iniciou'
            }

        sent = sum(1 for r in result.data if r['status'] == 'sent')
        failed = sum(1 for r in result.data if r['status'] == 'failed')
        total = len(result.data)

        return {
            'campaign_id': campaign_id,
            'found': True,
            'total_processed': total,
            'sent': sent,
            'failed': failed,
            'progress_percent': round((total / max(total, 1)) * 100, 1)
        }

    except Exception as e:
        return {
            'campaign_id': campaign_id,
            'found': False,
            'error': str(e)
        }


@router.get("/template")
def get_default_template():
    """Get the default reactivation message templates (2 messages)"""
    example_lead = {
        'name': 'Carlos',
        'notes': 'conseguir mais clientes pelo Google',
        'company': 'Auto Pecas Silva'
    }
    return {
        'msg1_template': DEFAULT_MSG1_TEMPLATE,
        'msg2_template': DEFAULT_MSG2_TEMPLATE,
        'placeholders': ['[NAME]', '[NOTES]', '[COMPANY]'],
        'example': {
            'msg1': format_message(DEFAULT_MSG1_TEMPLATE, example_lead),
            'msg2': format_message(DEFAULT_MSG2_TEMPLATE, example_lead)
        },
        'info': 'Envia 2 mensagens: (1) abertura/pergunta, (2) proposta + link. Sem delay entre elas.'
    }
